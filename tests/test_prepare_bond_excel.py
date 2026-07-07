import json
import sys
import tempfile
import unittest
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parents[1] / "skills" / "download-private-bond" / "scripts"
sys.path.insert(0, str(SCRIPT_DIR))

from prepare_bond_excel import (
    build_rows_json_path,
    fill_template,
    is_wind_value_ready,
    load_ocr_records,
    normalize_name,
    read_value_rows,
    remove_public_and_sort_rows,
    write_rows_json,
    write_final_rows,
)
import prepare_bond_excel
from openpyxl import Workbook, load_workbook


class PrepareBondExcelTests(unittest.TestCase):
    def test_load_ocr_records_accepts_bond_short_name_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bond_list.json"
            path.write_text(
                json.dumps(
                    [
                        {
                            "bond_short_name": "26嘉州01",
                        }
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            records = load_ocr_records(path)

        self.assertEqual(records[0]["company_name"], "")
        self.assertEqual(records[0]["bond_short_name"], "26嘉州01")

    def test_fill_template_writes_short_name_to_a_and_copies_formulas_b_to_e(self):
        with tempfile.TemporaryDirectory() as tmp:
            template_path = Path(tmp) / "template.xlsx"
            output_path = Path(tmp) / "output.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "债券简称"
            sheet["B1"] = "公司名称"
            sheet["C1"] = "债券代码"
            sheet["D1"] = "债券全称"
            sheet["E1"] = "发行方式"
            sheet["A2"] = "模板简称"
            sheet["B2"] = '=WSS(A2,"issuer")'
            sheet["C2"] = '=WSS(A2,"code")'
            sheet["D2"] = '=WSS(A2,"fullname")'
            sheet["E2"] = '=WSS(A2,"issue_method")'
            workbook.save(template_path)

            old_win32com = prepare_bond_excel.win32com
            prepare_bond_excel.win32com = None
            try:
                fill_template(
                    template_path,
                    output_path,
                    [{"company_name": "", "bond_short_name": "26嘉州01"}],
                )
            finally:
                prepare_bond_excel.win32com = old_win32com

            result = load_workbook(output_path, data_only=False).active

        self.assertEqual(result["A2"].value, "26嘉州01")
        self.assertEqual(result["B2"].value, '=WSS(A2,"issuer")')
        self.assertEqual(result["C2"].value, '=WSS(A2,"code")')
        self.assertEqual(result["D2"].value, '=WSS(A2,"fullname")')
        self.assertEqual(result["E2"].value, '=WSS(A2,"issue_method")')

    def test_read_and_write_value_rows_use_new_column_order(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel_path = Path(tmp) / "rows.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["债券简称", "公司名称", "债券代码", "债券全称", "发行方式"])
            sheet.append(["26嘉州01", "乐山国资", "CODE1", "乐山国资2026年非公开发行债券", "私募"])
            workbook.save(excel_path)

            rows = read_value_rows(excel_path)
            write_final_rows(excel_path, rows)
            result = load_workbook(excel_path).active

        self.assertEqual(rows[0]["bond_short_name"], "26嘉州01")
        self.assertEqual(rows[0]["company_name"], "乐山国资")
        self.assertEqual(result["A2"].value, "26嘉州01")
        self.assertEqual(result["B2"].value, "乐山国资")

    def test_wind_error_code_is_not_ready(self):
        self.assertFalse(is_wind_value_ready(-2146826259))
        self.assertFalse(is_wind_value_ready("Fetching..."))
        self.assertFalse(is_wind_value_ready(None))
        self.assertTrue(is_wind_value_ready("乐山国资"))

    def test_write_rows_json_uses_same_stem_as_excel(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel_path = Path(tmp) / "信评需求私募债_20260606_133633.xlsx"
            rows = [
                {
                    "company_name": "乐山国资",
                    "bond_short_name": "26嘉州01",
                    "bond_code": "282882.SH",
                    "bond_full_name": "乐山国资2026年非公开发行债券",
                    "issue_method": "私募",
                }
            ]

            json_path = write_rows_json(build_rows_json_path(excel_path), rows)
            data = json.loads(json_path.read_text(encoding="utf-8"))

        self.assertEqual(json_path.name, "信评需求私募债_20260606_133633.json")
        self.assertEqual(data[0]["bond_short_name"], "26嘉州01")
        self.assertEqual(data[0]["company_name"], "乐山国资")

    def test_normalize_name_unifies_parentheses_and_spaces(self):
        self.assertEqual(
            normalize_name(" 乐山国有资产投资运营（集团）有限公司 "),
            "乐山国有资产投资运营(集团)有限公司",
        )

    def test_remove_public_and_sort_rows(self):
        rows = [
            {
                "company_name": "乙公司",
                "bond_short_name": "26乙债01",
                "bond_code": "2",
                "bond_full_name": "乙公司2026年非公开发行债券",
                "issue_method": "私募",
            },
            {
                "company_name": "甲公司",
                "bond_short_name": "26甲债01",
                "bond_code": "1",
                "bond_full_name": "甲公司2026年公开发行债券",
                "issue_method": "公募",
            },
            {
                "company_name": "甲公司",
                "bond_short_name": "26甲私01",
                "bond_code": "3",
                "bond_full_name": "甲公司2026年非公开发行债券",
                "issue_method": "私募",
            },
        ]

        result = remove_public_and_sort_rows(rows)

        self.assertEqual([row["company_name"] for row in result], ["甲公司", "乙公司"])
        self.assertEqual([row["bond_short_name"] for row in result], ["26甲私01", "26乙债01"])


if __name__ == "__main__":
    unittest.main()

import argparse
import os
import re
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent if SCRIPT_DIR.name == "scripts" else SCRIPT_DIR


def find_project_root(start_path):
    for path in [start_path, *start_path.parents]:
        if (path / "data").exists() and (path / "reference").exists():
            return path
    return Path.cwd()


ROOT = find_project_root(SCRIPT_DIR)
DEFAULT_DOWNLOAD_DIR = ROOT / "Download"
DEFAULT_OUTPUT_DIR = ROOT / "output"
DEFAULT_EXCEL_PATTERN = "信评需求私募债_*.xlsx"
DEFAULT_WAIT_SECONDS = 30
INTERACTION_PAUSE_SECONDS = 0.5
TENANT_ANNOUNCEMENTS_URL = "https://www.ratingdog.cn/information/announcementsForTenant"


def load_skill_env(env_path=SKILL_DIR / ".env"):
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip().lstrip(":")
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def _import_selenium():
    from selenium import webdriver
    from selenium.common.exceptions import TimeoutException
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait

    return webdriver, TimeoutException, Options, By, EC, WebDriverWait


def placeholder_xpath(placeholder, input_type=None):
    xpath = f"//input[@placeholder='{placeholder}'"
    if input_type:
        xpath += f" and @type='{input_type}'"
    return xpath + "]"


def build_prospectus_search_keyword(bond_full_name):
    """构建募集说明书搜索关键词

    去掉债券名称中的 (品种X) 后缀，提高搜索匹配率
    例如：
    "杭州高新技术产业开发区资产经营有限公司2026年面向专业投资者非公开发行公司债券(第一期)(品种一)"
    → "杭州高新技术产业开发区资产经营有限公司2026年面向专业投资者非公开发行公司债券(第一期)募集说明书"
    """
    import re
    # 去掉 (品种X) 后缀，但保留其他括号内容如 (第一期)
    cleaned = re.sub(r'\(品种[一二三四五六七八九十]+\)', '', bond_full_name)
    return f"{cleaned}募集说明书"


def download_button_xpaths():
    # 先定位附件列(class包含column_12) → 找file-desc → 找包含"下载"文本的span
    return [
        ".//td[contains(@class,'column_12')]//span[@class='file-desc']/span[contains(text(),'下载')]",
        ".//td[contains(@class,'column_12')]//span[@class='file-desc']//span[contains(text(),'下载')]",
        ".//td[contains(@class,'column_12')]//span[contains(text(),'下载')]",
        ".//td[last()]//span[@class='file-desc']//span[contains(text(),'下载')]",
    ]


def search_button_xpaths():
    return [
        "./ancestor::*[contains(@class,'yyep-input-group')][1]//div[contains(@class,'yyep-input-group__append')]//button",
        "./ancestor::form[1]//div[contains(@class,'yyep-input-group__append')]//button",
        "./ancestor::form[1]//button[.//svg or contains(@class,'yyep-button')]",
    ]


def pause_between_interactions():
    time.sleep(INTERACTION_PAUSE_SECONDS)


def wait_visible(driver, by, locator, timeout=DEFAULT_WAIT_SECONDS):
    _, _, _, _, EC, WebDriverWait = _import_selenium()
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((by, locator))
    )


def wait_clickable(driver, by, locator, timeout=DEFAULT_WAIT_SECONDS):
    _, _, _, _, EC, WebDriverWait = _import_selenium()
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, locator))
    )


def click_when_ready(driver, by, locator, timeout=DEFAULT_WAIT_SECONDS):
    element = wait_clickable(driver, by, locator, timeout=timeout)
    driver.execute_script("arguments[0].click();", element)
    pause_between_interactions()
    return element


def input_when_ready(driver, by, locator, value, timeout=DEFAULT_WAIT_SECONDS):
    element = wait_visible(driver, by, locator, timeout=timeout)
    element.clear()
    if value:
        element.send_keys(value)
    pause_between_interactions()
    return element


def input_first_visible_when_ready(driver, locators, value, timeout=DEFAULT_WAIT_SECONDS):
    _, _, _, _, _, WebDriverWait = _import_selenium()
    element = WebDriverWait(driver, timeout).until(
        lambda _: first_visible_locator(driver, locators)
    )
    element.clear()
    if value:
        element.send_keys(value)
    pause_between_interactions()
    return element


def first_visible_locator(driver, locators):
    for by, locator in locators:
        try:
            element = first_displayed(driver.find_elements(by, locator))
            if element is not None:
                return element
        except Exception:
            continue
    return False


def first_displayed(elements):
    for element in elements:
        try:
            if element.is_displayed():
                return element
        except Exception:
            continue
    return None


def wait_child_visible(driver, parent, by, locator, timeout=DEFAULT_WAIT_SECONDS):
    _, _, _, _, _, WebDriverWait = _import_selenium()
    return WebDriverWait(driver, timeout).until(
        lambda _: safe_first_displayed(parent, by, locator)
    )


def click_child_when_ready(driver, parent, by, locator, timeout=DEFAULT_WAIT_SECONDS):
    element = wait_child_visible(driver, parent, by, locator, timeout=timeout)
    driver.execute_script("arguments[0].click();", element)
    pause_between_interactions()
    return element


def wait_existing_ready(driver, element, timeout=DEFAULT_WAIT_SECONDS):
    _, _, _, _, _, WebDriverWait = _import_selenium()
    WebDriverWait(driver, timeout).until(
        lambda _: element_is_ready(element)
    )
    return element


def click_element_when_ready(driver, element, timeout=DEFAULT_WAIT_SECONDS):
    wait_existing_ready(driver, element, timeout=timeout)
    driver.execute_script("arguments[0].click();", element)
    pause_between_interactions()
    return element


def send_keys_when_ready(driver, element, *keys, timeout=DEFAULT_WAIT_SECONDS):
    wait_existing_ready(driver, element, timeout=timeout)
    element.send_keys(*keys)
    pause_between_interactions()
    return element


def safe_first_displayed(parent, by, locator):
    try:
        return first_displayed(parent.find_elements(by, locator))
    except Exception:
        return False


def element_is_ready(element):
    try:
        return element.is_displayed() and element.is_enabled()
    except Exception:
        return False


def extract_year_from_bond_full_name(value):
    match = re.search(r"(20\d{2})年", str(value))
    if not match:
        raise ValueError(f"债券全称中未找到年份：{value}")
    return int(match.group(1))


def normalize_for_match(value):
    text = str(value).strip()
    replacements = {
        "（": "(",
        "）": ")",
        "“": "",
        "”": "",
        "\"": "",
        " ": "",
        "\u3000": "",
        "\n": "",
        "\t": "",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def title_matches_bond(title, bond_full_name):
    return normalize_for_match(bond_full_name) in normalize_for_match(title)


def is_target_prospectus_title(title, bond_full_name):
    """判断是否为目标的募集说明书标题

    匹配条件：
    1. 标题包含"募集说明书"
    2. 标题包含债券全称（去掉品种后缀后）
    """
    import re
    if "募集说明书" not in str(title):
        return False

    # 去掉债券全称中的 (品种X) 后缀
    cleaned_bond = re.sub(r'\(品种[一二三四五六七八九十]+\)', '', bond_full_name)

    # 标准化后比较
    normalized_title = normalize_for_match(title)
    normalized_bond = normalize_for_match(cleaned_bond)

    return normalized_bond in normalized_title


def group_consecutive_issuers(rows):
    groups = []
    current = []
    current_company = None
    for row in rows:
        company = row["company_name"]
        if current and company != current_company:
            groups.append(current)
            current = []
        current.append(row)
        current_company = company
    if current:
        groups.append(current)
    return groups


def read_bond_rows(excel_path):
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active
    rows = []
    for row_index in range(2, sheet.max_row + 1):
        company_name = sheet.cell(row=row_index, column=1).value
        bond_short_name = sheet.cell(row=row_index, column=2).value
        bond_code = sheet.cell(row=row_index, column=3).value
        bond_full_name = sheet.cell(row=row_index, column=4).value
        issue_method = sheet.cell(row=row_index, column=5).value
        if not company_name or not bond_full_name:
            continue
        rows.append(
            {
                "company_name": str(company_name).strip().replace("（", "(").replace("）", ")"),
                "bond_short_name": str(bond_short_name or "").strip(),
                "bond_code": str(bond_code or "").strip(),
                "bond_full_name": str(bond_full_name).strip(),
                "issue_method": str(issue_method or "").strip(),
            }
        )
    return rows


def find_latest_prepared_excel(output_dir=DEFAULT_OUTPUT_DIR):
    output_dir = Path(output_dir)
    candidates = [
        path for path in output_dir.glob(DEFAULT_EXCEL_PATTERN)
        if path.is_file()
    ]
    if not candidates:
        raise FileNotFoundError(
            f"未在 {output_dir} 找到 {DEFAULT_EXCEL_PATTERN}，请先运行 prepare_bond_excel.py 或手动传入 --excel"
        )
    return max(candidates, key=lambda path: path.stat().st_mtime)


def resolve_excel_path(excel_path, output_dir=DEFAULT_OUTPUT_DIR):
    if excel_path:
        return Path(excel_path)
    return find_latest_prepared_excel(output_dir)


def create_driver(download_dir):
    webdriver, _, Options, _, _, _ = _import_selenium()
    download_dir = Path(download_dir)
    download_dir.mkdir(parents=True, exist_ok=True)
    prefs = {
        "download.default_directory": str(download_dir.resolve()),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "safebrowsing.enabled": False,
    }
    options = Options()
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(options=options)
    driver.download_dir = download_dir
    driver.set_window_size(1400, 900)
    return driver


def find_visible(driver, by, locator, timeout=20):
    return wait_visible(driver, by, locator, timeout=timeout)


def visible_input(driver, placeholder, timeout=20):
    _, _, _, By, _, _ = _import_selenium()
    return wait_visible(driver, By.XPATH, placeholder_xpath(placeholder), timeout=timeout)


def click_if_visible(driver, xpath, timeout=3):
    _, TimeoutException, _, By, _, _ = _import_selenium()
    try:
        click_when_ready(driver, By.XPATH, xpath, timeout=timeout)
        return True
    except TimeoutException:
        return False


def login(driver, username, password):
    _, _, _, By, _, _ = _import_selenium()

    print("[DEBUG] 正在打开登录页面...")
    driver.get("https://www.ratingdog.cn/login")
    print(f"[DEBUG] 当前URL: {driver.current_url}")

    print("[DEBUG] 切换到密码登录标签...")
    click_when_ready(driver, By.ID, "tab-1")
    print("[DEBUG] 输入用户名...")
    input_first_visible_when_ready(
        driver,
        [
            (By.CSS_SELECTOR, "#pane-2 input[placeholder='请输入']"),
            (By.XPATH, placeholder_xpath("请输入手机号码")),
        ],
        username,
    )
    print("[DEBUG] 输入密码...")
    input_first_visible_when_ready(
        driver,
        [
            (By.CSS_SELECTOR, "#pane-2 input[placeholder='请输入密码']"),
            (By.XPATH, placeholder_xpath("请输入密码", input_type="password")),
        ],
        password,
    )
    print("[DEBUG] 点击登录按钮...")
    click_when_ready(driver, By.CSS_SELECTOR, "button.yyep-button--primary")
    print("[DEBUG] 等待登录成功...")
    wait_visible(driver, By.XPATH, "//span[@class='txtColor' and text()='公告信息']")
    print("[DEBUG] 登录成功，正在跳转公告信息页面...")
    driver.get(TENANT_ANNOUNCEMENTS_URL)
    print(f"[DEBUG] 已跳转至: {driver.current_url}")
    wait_visible(driver, By.XPATH, placeholder_xpath("标题、简称、发行人、债券代码"))
    print("[DEBUG] 公告信息页面加载完成")
    #driver.get(TENANT_ANNOUNCEMENTS_URL)
    #wait_visible(driver, By.XPATH, placeholder_xpath("标题、简称、发行人、债券代码"))

    





def safe_attribute(element, name):
    try:
        return element.get_attribute(name) or ""
    except Exception:
        return ""


def login_state_ready(driver):
    if "login" not in driver.current_url:
        return True
    try:
        return bool(
            driver.execute_script(
                """
                const keys = Object.keys(window.localStorage || {});
                return keys.some(key => /token|auth/i.test(key) && localStorage.getItem(key));
                """
            )
        )
    except Exception:
        return False


def _has_login_state(driver):
    if "login" not in driver.current_url:
        return True
    try:
        return bool(
            driver.execute_script(
                """
                const keys = Object.keys(window.localStorage || {});
                return keys.some(key => /token|auth|user/i.test(key) && localStorage.getItem(key));
                """
            )
        )
    except Exception:
        return False


def dismiss_date_picker(driver, fallback_element=None):
    from selenium.webdriver.common.keys import Keys

    _, _, _, By, _, _ = _import_selenium()
    active_element = driver.switch_to.active_element
    send_keys_when_ready(driver, active_element, Keys.ENTER)
    if wait_date_picker_closed(driver, timeout=3):
        return

    if click_date_picker_confirm(driver):
        wait_date_picker_closed(driver)
        return

    send_keys_when_ready(driver, active_element, Keys.ESCAPE)
    if wait_date_picker_closed(driver, timeout=3):
        return

    click_when_ready(driver, By.TAG_NAME, "body")
    wait_date_picker_closed(driver, timeout=3)


def click_date_picker_confirm(driver):
    _, TimeoutException, _, By, _, _ = _import_selenium()
    confirm_xpaths = [
        "//div[contains(@class,'yyep-picker-panel') and not(contains(@style,'display: none'))]//button[normalize-space()='确定']",
        "//div[contains(@class,'yyep-picker-panel') and not(contains(@style,'display: none'))]//button[normalize-space()='确认']",
        "//div[contains(@class,'yyep-picker-panel') and not(contains(@style,'display: none'))]//button[contains(@class,'yyep-picker-panel__link-btn') and not(@disabled)]",
    ]
    for xpath in confirm_xpaths:
        try:
            click_when_ready(driver, By.XPATH, xpath, timeout=3)
            return True
        except TimeoutException:
            continue
    return False


def wait_date_picker_closed(driver, timeout=DEFAULT_WAIT_SECONDS):
    _, TimeoutException, _, _, _, WebDriverWait = _import_selenium()
    try:
        WebDriverWait(driver, timeout).until(
            lambda current_driver: not current_driver.execute_script(
                """
                return [...document.querySelectorAll('.yyep-picker-panel')]
                  .some(el => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                """
            )
        )
        return True
    except TimeoutException:
        return False


def set_date_range(driver, start_input, end_input, year):
    from selenium.webdriver.common.keys import Keys

    start_value = f"{year}-01-01" if year else ""
    end_value = f"{year}-12-31" if year else ""

    def set_input_value(element, value):
        click_element_when_ready(driver, element)
        send_keys_when_ready(driver, element, Keys.CONTROL, "a")
        send_keys_when_ready(driver, element, Keys.BACKSPACE)
        if value:
            send_keys_when_ready(driver, element, value)
        driver.execute_script(
            """
            const el = arguments[0];
            const value = arguments[1];
            const setter = Object.getOwnPropertyDescriptor(
              window.HTMLInputElement.prototype,
              'value'
            ).set;
            setter.call(el, value);
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
            el.dispatchEvent(new Event('blur', { bubbles: true }));
            """,
            element,
            value,
        )

    set_input_value(start_input, start_value)
    set_input_value(end_input, end_value)
    dismiss_date_picker(driver, fallback_element=start_input)

    actual_start = start_input.get_attribute("value") or ""
    actual_end = end_input.get_attribute("value") or ""
    if start_value and actual_start != start_value:
        driver.execute_script(
            """
            const el = arguments[0];
            el.value = arguments[1];
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
            """,
            start_input,
            start_value,
        )
    if end_value and actual_end != end_value:
        driver.execute_script(
            """
            const el = arguments[0];
            el.value = arguments[1];
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
            """,
            end_input,
            end_value,
        )
    wait_date_picker_closed(driver, timeout=3)


def wait_for_search_results(driver, bond_full_name, previous_first_title="", timeout=DEFAULT_WAIT_SECONDS):
    """等待搜索结果加载完成

    判断依据：
    1. 表格行数 >= 1 且 < 20
    2. 第一行标题与上一只债券不同（数据已刷新）
    3. 第一行标题包含当前债券全称
    4. 行数稳定（连续2次检测相同）
    5. 或出现空数据提示
    """
    _, _, _, By, _, _ = _import_selenium()
    deadline = time.time() + timeout
    last_row_count = 0
    stable_count = 0

    while time.time() < deadline:
        # 检查可见行数（优先）
        rows = visible_table_rows(driver)
        current_count = len(rows)

        # 如果有数据行，不再检查空数据提示（优先使用表格数据）
        if current_count >= 1 and current_count < 20:
            try:
                title_cells = rows[0].find_elements(By.XPATH, "./td[1]")
                if title_cells:
                    first_title = title_cells[0].text.strip()
                    normalized_title = normalize_for_match(first_title)

                    # 如果不是第一只债券，先检查是否与上一只不同（数据已刷新）
                    if previous_first_title:
                        normalized_prev = normalize_for_match(previous_first_title)
                        if normalized_title == normalized_prev:
                            print(f"[DEBUG] 数据未刷新，仍是上一只债券的结果: {first_title[:50]}...")
                            stable_count = 0
                            last_row_count = current_count
                            time.sleep(0.5)
                            continue

                    # 检查是否包含当前债券全称（使用募集说明书匹配逻辑）
                    if is_target_prospectus_title(first_title, bond_full_name):
                        if current_count == last_row_count:
                            stable_count += 1
                            if stable_count >= 2:
                                print(f"[DEBUG] 搜索结果加载完成，共 {current_count} 行，第一行匹配")
                                return rows
                        else:
                            stable_count = 0
                            print(f"[DEBUG] 搜索结果变化中: {last_row_count} -> {current_count}")
                    else:
                        print(f"[DEBUG] 第一行不匹配当前债券: {first_title[:50]}...")
            except Exception as e:
                print(f"[DEBUG] 检查第一行时出错: {e}")

        # 只有在没有数据行时，才检查空数据提示
        elif current_count == 0 and has_empty_result_hint(driver):
            print("[DEBUG] 检测到空数据提示")
            return []

        last_row_count = current_count
        time.sleep(0.5)

    # 超时返回当前可见行
    final_rows = visible_table_rows(driver)
    print(f"[DEBUG] 等待超时，返回当前 {len(final_rows)} 行")
    return final_rows


def search_tenant_announcements(driver, bond_full_name, year, is_first=False, previous_first_title=""):
    _, _, _, By, _, _ = _import_selenium()
    print(f"[DEBUG] 开始搜索债券: {bond_full_name}, 年份: {year}, 是否首个: {is_first}")

    print("[DEBUG] 定位开始日期输入框...")
    start_input = visible_input(driver, "开始日期")
    print("[DEBUG] 定位结束日期输入框...")
    end_input = visible_input(driver, "结束日期")
    print(f"[DEBUG] 设置日期范围: {year}-01-01 至 {year}-12-31")
    set_date_range(driver, start_input, end_input, year)
    print("[DEBUG] 日期范围设置完成")

    keyword = build_prospectus_search_keyword(bond_full_name)
    print(f"[DEBUG] 搜索关键词: {keyword}")
    search_input = input_when_ready(
        driver,
        By.XPATH,
        placeholder_xpath("标题、简称、发行人、债券代码"),
        keyword,
    )

    print("[DEBUG] 点击搜索按钮...")
    click_search_button_for_input(driver, search_input)

    # 第一个债券特殊处理：3秒后再次点击搜索（避免页面自动刷新全量公告）
    if is_first:
        print("[DEBUG] 首个债券，等待3秒后再次搜索...")
        time.sleep(3)
        print("[DEBUG] 再次点击搜索按钮...")
        click_search_button_for_input(driver, search_input)

    print("[DEBUG] 等待搜索结果加载...")
    rows = wait_for_search_results(driver, bond_full_name, previous_first_title=previous_first_title, timeout=DEFAULT_WAIT_SECONDS)

    print(f"[DEBUG] 搜索结果行数: {len(rows)}")
    return rows


def click_search_button_for_input(driver, search_input):
    _, _, _, By, _, _ = _import_selenium()
    for xpath in search_button_xpaths():
        try:
            click_child_when_ready(driver, search_input, By.XPATH, xpath, timeout=5)
            return
        except Exception:
            continue
    raise RuntimeError("未找到搜索按钮")


def wait_for_results_count_below(driver, limit, timeout=DEFAULT_WAIT_SECONDS):
    deadline = time.time() + timeout
    while time.time() < deadline:
        total_count = read_pagination_total_count(driver)
        if total_count is not None and total_count < limit:
            return visible_table_rows(driver)
        if has_empty_result_hint(driver):
            return []
        time.sleep(0.5)
    return visible_table_rows(driver)


def read_pagination_total_count(driver):
    _, _, _, By, _, _ = _import_selenium()
    totals = driver.find_elements(By.CSS_SELECTOR, "span.yyep-pagination__total")
    for total in totals:
        try:
            if not total.is_displayed():
                continue
            match = re.search(r"共\s*(\d+)\s*条", total.text)
            if match:
                return int(match.group(1))
        except Exception:
            continue
    return None


def visible_table_rows(driver):
    _, _, _, By, _, _ = _import_selenium()
    rows = []
    for row in driver.find_elements(By.XPATH, "//tbody/tr"):
        try:
            if row.is_displayed():
                rows.append(row)
        except Exception:
            continue
    return rows


def has_empty_result_hint(driver):
    _, _, _, By, _, _ = _import_selenium()
    empty_hints = driver.find_elements(By.XPATH, "//*[contains(text(),'暂无数据') or contains(text(),'无数据')]")
    for element in empty_hints:
        try:
            if element.is_displayed():
                return True
        except Exception:
            continue
    return False


def wait_for_matching_announcement_or_empty(driver, bond_full_name, timeout=DEFAULT_WAIT_SECONDS):
    _, TimeoutException, _, _, _, _ = _import_selenium()
    deadline = time.time() + timeout
    while time.time() < deadline:
        for row in visible_table_rows(driver):
            try:
                title_cells = row.find_elements(By.XPATH, "./td[1]")
                if title_cells and title_matches_bond(title_cells[0].text.strip(), bond_full_name):
                    return
            except Exception:
                continue
        if has_empty_result_hint(driver):
            return
        time.sleep(0.2)
    raise TimeoutException("公告搜索后未出现匹配标题或空结果提示")


def wait_for_announcement_results(driver, timeout=30):
    _, TimeoutException, _, By, _, _ = _import_selenium()
    deadline = time.time() + timeout
    while time.time() < deadline:
        rows = driver.find_elements(By.XPATH, "//tbody/tr")
        for row in rows:
            try:
                if row.is_displayed():
                    return
            except Exception:
                continue
        empty_hints = driver.find_elements(By.XPATH, "//*[contains(text(),'暂无数据') or contains(text(),'无数据')]")
        for element in empty_hints:
            try:
                if element.is_displayed():
                    return
            except Exception:
                continue
        time.sleep(0.2)
    raise TimeoutException("公告搜索后未出现结果行或空结果提示")


def close_download_dialog(driver):
    _, TimeoutException, _, By, _, _ = _import_selenium()
    try:
        click_when_ready(
            driver,
            By.XPATH,
            "//*[@role='dialog' and not(contains(@style,'display: none'))]//button[.//*[text()='确定']]",
        )
    except TimeoutException:
        pass


def find_and_download(driver, bond_full_name, rows):
    """在已搜索到的行中查找并下载募集说明书

    Args:
        driver: Selenium driver
        bond_full_name: 债券全称
        rows: 搜索返回的结果行（避免重新获取导致错位）
    """
    _, _, _, By, _, _ = _import_selenium()
    print(f"[DEBUG] 开始在 {len(rows)} 行中查找募集说明书: {bond_full_name}")

    for attempt in range(3):
        print(f"[DEBUG] 第 {attempt + 1} 次尝试查找...")

        # 使用传入的rows，不再重新获取
        if attempt > 0:
            # 重试时重新获取（可能页面有变化）
            rows = visible_table_rows(driver)
            print(f"[DEBUG] 重试，当前可见行数: {len(rows)}")

        for i, row in enumerate(rows):
            try:
                title_cells = row.find_elements(By.XPATH, "./td[1]")
                if not title_cells:
                    continue
                title_text = title_cells[0].text.strip()
                print(f"[DEBUG] 第 {i+1} 行标题: {title_text[:50]}...")

                if is_target_prospectus_title(title_text, bond_full_name):
                    print(f"[DEBUG] 找到匹配标题，准备下载...")
                    return click_row_download(driver, row, title_text)
            except Exception as e:
                print(f"[DEBUG] 处理第 {i+1} 行时出错: {e}")
                continue

        # 未找到匹配，尝试第一行
        if rows:
            try:
                title_cells = rows[0].find_elements(By.XPATH, "./td[1]")
                title_text = title_cells[0].text.strip() if title_cells else "第一条搜索结果"
                print(f"[DEBUG] 未找到匹配标题，尝试下载第一行: {title_text[:50]}...")
                return click_row_download(driver, rows[0], title_text)
            except Exception as e:
                print(f"[DEBUG] 尝试第一行下载失败: {e}")
        time.sleep(0.3)

    return False, "未找到标题匹配的募集说明书公告"


def click_row_download(driver, row, title_text):
    """点击行内的下载按钮，找不到直接报错"""
    _, _, _, By, _, _ = _import_selenium()
    # 使用最可靠的XPath：最后一列任何包含"下载"的元素
    xpath = ".//td[last()]//*[contains(text(),'下载')]"
    print(f"[DEBUG] 使用XPath查找下载按钮: {xpath}")
    download_elements = row.find_elements(By.XPATH, xpath)
    print(f"[DEBUG] 找到 {len(download_elements)} 个下载按钮元素")

    if not download_elements:
        # 打印所有td的class帮助调试
        try:
            tds = row.find_elements(By.XPATH, "./td")
            print(f"[DEBUG] 行内共有 {len(tds)} 列")
            for i, td in enumerate(tds):
                class_attr = td.get_attribute('class')
                print(f"[DEBUG] 第{i+1}列 class: {class_attr}")
                text = td.text[:100] if td.text else "空"
                print(f"[DEBUG] 第{i+1}列内容: {text}")
        except Exception as e:
            print(f"[DEBUG] 无法获取列信息: {e}")
        raise RuntimeError(f"未找到下载按钮：{title_text}")

    print(f"[DEBUG] 准备点击下载按钮...")
    before_files = snapshot_download_files(driver)
    print(f"[DEBUG] 下载前文件列表: {before_files}")

    click_element_when_ready(driver, download_elements[0])
    print(f"[DEBUG] 已点击下载按钮，等待下载完成...")

    close_download_dialog(driver)
    downloaded = wait_for_download_complete(driver, before_files)

    if downloaded:
        print(f"[DEBUG] 下载成功: {downloaded}")
        return True, title_text
    else:
        print(f"[DEBUG] 下载超时，未检测到新文件")
        return False, f"下载未在限定时间内完成：{title_text}"


def snapshot_download_files(driver):
    download_dir = Path(getattr(driver, "download_dir", DEFAULT_DOWNLOAD_DIR))
    download_dir.mkdir(parents=True, exist_ok=True)
    return {path.name for path in download_dir.iterdir() if path.is_file()}


def wait_for_download_complete(driver, before_files, timeout=120):
    download_dir = Path(getattr(driver, "download_dir", DEFAULT_DOWNLOAD_DIR))
    print(f"[DEBUG] 等待下载完成，目录: {download_dir}, 超时: {timeout}s")
    deadline = time.time() + timeout
    check_count = 0
    while time.time() < deadline:
        check_count += 1
        files = [path for path in download_dir.iterdir() if path.is_file()]
        new_files = [path for path in files if path.name not in before_files]
        partial_files = [path for path in new_files if path.name.endswith(".crdownload")]
        complete_pdfs = [path for path in new_files if path.suffix.lower() == ".pdf"]

        if check_count % 10 == 0:  # 每5秒打印一次
            print(f"[DEBUG] 检查 #{check_count}: 新文件 {len(new_files)}, 部分下载 {len(partial_files)}, 完成PDF {len(complete_pdfs)}")

        if complete_pdfs and not partial_files:
            print(f"[DEBUG] 下载完成: {complete_pdfs[0].name}")
            return complete_pdfs[0]
        time.sleep(0.5)

    print(f"[DEBUG] 下载超时，未检测到完成文件")
    return None


def write_log(log_path, lines):
    Path(log_path).parent.mkdir(parents=True, exist_ok=True)
    Path(log_path).write_text("\n".join(lines) + "\n", encoding="utf-8")


def process_rows(driver, rows):
    log_lines = []
    valid_rows = []
    print(f"[DEBUG] 开始处理 {len(rows)} 行数据")

    for row in rows:
        try:
            extract_year_from_bond_full_name(row["bond_full_name"])
            valid_rows.append(row)
        except ValueError as exc:
            print(f"[DEBUG] 跳过行(无年份): {row.get('company_name')} - {exc}")
            log_lines.append(
                f"{row['company_name']}\t{row['bond_short_name']}\t{row['bond_full_name']}\t{exc}"
            )

    print(f"[DEBUG] 有效行数: {len(valid_rows)}")

    previous_first_title = ""  # 记录上一只债券的第一行标题

    for idx, row in enumerate(valid_rows):
        company_name = row["company_name"]
        bond_full_name = row["bond_full_name"]
        print(f"\n[DEBUG] === 处理第 {idx+1}/{len(valid_rows)} 只债券 ===")
        print(f"[DEBUG] 发行人: {company_name}")
        print(f"[DEBUG] 债券全称: {bond_full_name}")

        try:
            year = extract_year_from_bond_full_name(bond_full_name)
            print(f"[DEBUG] 提取年份: {year}")

            rows = search_tenant_announcements(driver, bond_full_name, year, is_first=(idx==0), previous_first_title=previous_first_title)
            ok, message = find_and_download(driver, bond_full_name, rows)

            # 记录本次的第一行标题，用于下一次比较
            if rows:
                try:
                    from selenium.webdriver.common.by import By
                    title_cells = rows[0].find_elements(By.XPATH, "./td[1]")
                    if title_cells:
                        previous_first_title = title_cells[0].text.strip()
                        print(f"[DEBUG] 记录第一行标题供下次比较: {previous_first_title[:50]}...")
                except Exception as e:
                    print(f"[DEBUG] 记录第一行标题失败: {e}")

            if ok:
                print(f"[DEBUG] 下载成功: {message}")
            else:
                print(f"[DEBUG] 下载失败: {message}")
                log_lines.append(
                    f"{company_name}\t{row['bond_short_name']}\t{bond_full_name}\t{message}"
                )
        except Exception as e:
            print(f"[DEBUG] 处理异常: {e}")
            log_lines.append(
                f"{company_name}\t{row['bond_short_name']}\t{bond_full_name}\t{str(e)}"
            )

    print(f"\n[DEBUG] 全部处理完成，失败记录: {len(log_lines)}")
    # 返回日志行和失败的债券记录（用于第二轮搜索）
    failed_bonds = []
    for line in log_lines:
        parts = line.split('\t')
        if len(parts) >= 3:
            failed_bonds.append({
                'company_name': parts[0],
                'bond_short_name': parts[1],
                'bond_full_name': parts[2],
                'error': parts[3] if len(parts) > 3 else '未知错误'
            })
    return log_lines, failed_bonds


def main():
    load_skill_env()
    parser = argparse.ArgumentParser(description="从 Ratingdog 下载私募债相关公告附件")
    parser.add_argument("--excel", required=True, help="prepare_bond_excel.py 输出的 Excel")
    parser.add_argument("--download-dir", default=str(DEFAULT_DOWNLOAD_DIR), help="PDF 下载目录")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="日志输出目录")
    parser.add_argument("--username", default=os.environ.get("RATINGDOG_USERNAME", ""), help="Ratingdog 手机号")
    parser.add_argument("--password", default=os.environ.get("RATINGDOG_PASSWORD", ""), help="Ratingdog 密码")
    for action in parser._actions:
        if action.dest == "excel":
            action.required = False
            action.help = "prepare_bond_excel.py 输出的 Excel；未提供时自动使用 output 目录下最新的 信评需求私募债_*.xlsx"
            break
    args = parser.parse_args()

    print(f"[DEBUG] 下载目录: {args.download_dir}")
    print(f"[DEBUG] 输出目录: {args.output_dir}")
    print(f"[DEBUG] 用户名: {args.username[:3]}****" if args.username else "[DEBUG] 用户名: 未设置")

    if not args.username or not args.password:
        raise ValueError("请通过参数或环境变量提供 RATINGDOG_USERNAME/RATINGDOG_PASSWORD")

    excel_path = resolve_excel_path(args.excel, args.output_dir)
    print(f"Excel：{excel_path}")

    rows = read_bond_rows(excel_path)
    print(f"[DEBUG] 从Excel读取 {len(rows)} 行数据")

    print("[DEBUG] 创建Chrome浏览器...")
    driver = create_driver(args.download_dir)
    print(f"[DEBUG] 浏览器下载目录: {driver.download_dir}")

    try:
        login(driver, args.username, args.password)
        log_lines, failed_bonds = process_rows(driver, rows)
    finally:
        print("[DEBUG] 关闭浏览器...")
        driver.quit()

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = Path(args.output_dir) / f"download_log_{stamp}.txt"
    write_log(log_path, log_lines or ["全部债券均已处理，未记录失败项"])
    print(f"日志：{log_path}")


if __name__ == "__main__":
    main()

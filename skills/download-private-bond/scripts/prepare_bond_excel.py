import argparse
import json
import locale
import shutil
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    import win32com.client
except ImportError:
    win32com = None


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent if SCRIPT_DIR.name == "scripts" else SCRIPT_DIR


def find_project_root(start_path):
    for path in [start_path, *start_path.parents]:
        if (path / "data").exists() and (path / "reference").exists():
            return path
    return Path.cwd()


ROOT = find_project_root(SCRIPT_DIR)
DEFAULT_INPUT = ROOT / "data" / "bond_list.json"
DEFAULT_TEMPLATE = SKILL_DIR / "assets" / "公式模板.xlsx"
DEFAULT_OUTPUT_DIR = ROOT / "output"


def _set_chinese_collation():
    for locale_name in ("Chinese_China.936", "zh_CN.UTF-8", "chs"):
        try:
            locale.setlocale(locale.LC_COLLATE, locale_name)
            return
        except locale.Error:
            continue


_set_chinese_collation()


def normalize_name(value):
    return str(value).strip().replace("（", "(").replace("）", ")")


def extract_company_from_bond_full_name(bond_full_name):
    """从债券全称中提取发行人名称

    债券全称格式：发行人名称2026年面向...
    提取逻辑：从头开始，找到第一个年份数字（20xx年）为止
    """
    if not bond_full_name:
        return ""

    import re
    match = re.search(r'(20\d{2}年)', str(bond_full_name))
    if match:
        company = str(bond_full_name)[:match.start()].strip()
        return company
    return ""


def load_ocr_records(path):
    """加载OCR识别结果

    支持两种模式：
    1. 完整模式：company_name + bond_short_name
    2. 简写模式：仅 bond_short_name（公司名称通过WIND公式自动填充）
    """
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    records = []
    for index, item in enumerate(data, start=1):
        company_name = normalize_name(item.get("company_name", ""))
        bond_short_name = str(item.get("bond_short_name", "")).strip()
        # 只要求债券简称必须存在，公司名称可以为空（由WIND公式自动填充）
        if not bond_short_name:
            raise ValueError(f"OCR记录第{index}行缺少债券简称")
        records.append(
            {
                "company_name": company_name,  # 可能为空，由WIND公式自动填充
                "bond_short_name": bond_short_name,
            }
        )
    return records


def remove_public_and_sort_rows(rows):
    private_rows = [
        row for row in rows
        if str(row.get("issue_method", "")).strip() != "公募"
    ]
    return sorted(
        private_rows,
        key=lambda row: (
            # 如果公司名称为空，使用债券简称排序
            locale.strxfrm(normalize_name(row.get("company_name", row.get("bond_short_name", "")))),
            str(row.get("bond_short_name", "")).strip(),
        ),
    )


def build_output_path(output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_dir / f"信评需求私募债_{stamp}.xlsx"


def _translate_template_formula(formula, row_index):
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    return formula.replace("B2", f"B{row_index}").replace("C2", f"C{row_index}")


def fill_template(template_path, output_path, records):
    shutil.copy2(template_path, output_path)
    if win32com is not None:
        return fill_template_with_excel(output_path, records)

    workbook = load_workbook(output_path)
    sheet = workbook.active

    formula_templates = {
        column: sheet.cell(row=2, column=column).value
        for column in range(3, 6)
    }

    for row_index, record in enumerate(records, start=2):
        # 如果公司名称不为空，则填入；为空则保留WIND公式自动填充
        if record["company_name"]:
            sheet.cell(row=row_index, column=1).value = record["company_name"]
        sheet.cell(row=row_index, column=2).value = record["bond_short_name"]
        for column, formula in formula_templates.items():
            sheet.cell(row=row_index, column=column).value = _translate_template_formula(formula, row_index)

    if sheet.max_row > len(records) + 1:
        for row_index in range(len(records) + 2, sheet.max_row + 1):
            for column in range(1, 6):
                sheet.cell(row=row_index, column=column).value = None

    workbook.save(output_path)
    return output_path


def fill_template_with_excel(output_path, records):
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = True
    excel.DisplayAlerts = False
    workbook = excel.Workbooks.Open(str(Path(output_path).resolve()), UpdateLinks=0)
    try:
        sheet = workbook.Worksheets(1)
        formula_templates = {
            column: sheet.Cells(2, column).Formula
            for column in range(3, 6)
        }
        used_rows = sheet.UsedRange.Rows.Count
        if used_rows >= 2:
            sheet.Range(f"A2:E{max(used_rows, len(records) + 1)}").ClearContents()

        for offset, record in enumerate(records):
            row_index = offset + 2
            # 如果公司名称不为空，则填入；为空则保留WIND公式自动填充
            if record["company_name"]:
                sheet.Cells(row_index, 1).Value = record["company_name"]
            sheet.Cells(row_index, 2).Value = record["bond_short_name"]
            for column, formula in formula_templates.items():
                if isinstance(formula, str) and formula.startswith("="):
                    sheet.Cells(row_index, column).Formula = _translate_template_formula(formula, row_index)

        workbook.Save()
    finally:
        workbook.Close(SaveChanges=True)
        excel.Quit()
    return output_path


def _flatten_excel_range(values):
    if values is None:
        return []
    if not isinstance(values, tuple):
        return [values]
    flat_values = []
    for row in values:
        if isinstance(row, tuple):
            flat_values.extend(row)
        else:
            flat_values.append(row)
    return flat_values


def refresh_wind_values(excel_path, max_wait=120):
    if win32com is None:
        raise RuntimeError("未安装 pywin32，无法通过 Excel COM 刷新 WIND 公式")

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    excel.DisplayAlerts = False
    workbook = excel.Workbooks.Open(str(Path(excel_path).resolve()), UpdateLinks=0)
    try:
        excel.CalculateFullRebuild()
        sheet = workbook.Worksheets(1)
        last_row = sheet.UsedRange.Rows.Count
        deadline = time.time() + max_wait

        while time.time() < deadline:
            values = sheet.Range(f"C2:E{last_row}").Value
            flat_values = _flatten_excel_range(values)
            if flat_values and all(value not in (None, "", "Fetching...") for value in flat_values):
                break
            time.sleep(2)
        else:
            raise TimeoutError("等待 WIND 公式计算超时")

        value_range = sheet.Range(f"C2:E{last_row}")
        value_range.Value = value_range.Value
        workbook.Save()
    finally:
        workbook.Close(SaveChanges=True)
        excel.Quit()


def read_value_rows(excel_path):
    workbook = load_workbook(excel_path, data_only=False)
    sheet = workbook.active
    rows = []
    for row_index in range(2, sheet.max_row + 1):
        company_name = sheet.cell(row=row_index, column=1).value
        bond_short_name = sheet.cell(row=row_index, column=2).value
        # 只要求债券简称存在，公司名称可以为空（由WIND公式自动填充）
        if not bond_short_name:
            continue
        rows.append(
            {
                "company_name": normalize_name(company_name) if company_name else "",
                "bond_short_name": str(bond_short_name).strip(),
                "bond_code": sheet.cell(row=row_index, column=3).value,
                "bond_full_name": sheet.cell(row=row_index, column=4).value,
                "issue_method": sheet.cell(row=row_index, column=5).value,
            }
        )
    return rows


def write_final_rows(excel_path, rows):
    workbook = load_workbook(excel_path)
    if hasattr(workbook, "_external_links"):
        workbook._external_links = []
    sheet = workbook.active
    if sheet.max_row >= 2:
        sheet.delete_rows(2, sheet.max_row - 1)
    for row_index, row in enumerate(rows, start=2):
        sheet.cell(row=row_index, column=1).value = row["company_name"]
        sheet.cell(row=row_index, column=2).value = row["bond_short_name"]
        sheet.cell(row=row_index, column=3).value = row["bond_code"]
        sheet.cell(row=row_index, column=4).value = row["bond_full_name"]
        sheet.cell(row=row_index, column=5).value = row["issue_method"]
    for column in range(1, 6):
        sheet.column_dimensions[get_column_letter(column)].width = 24
    workbook.save(excel_path)


def main():
    parser = argparse.ArgumentParser(description="生成并刷新私募债 WIND 数据 Excel")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="agent OCR 结果 JSON")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="公式模板 xlsx")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--skip-wind", action="store_true", help="跳过 WIND 刷新，仅用于开发测试")
    args = parser.parse_args()

    records = load_ocr_records(args.input)
    output_path = build_output_path(args.output_dir)
    fill_template(args.template, output_path, records)
    if not args.skip_wind:
        refresh_wind_values(output_path)
    final_rows = remove_public_and_sort_rows(read_value_rows(output_path))
    write_final_rows(output_path, final_rows)
    print(f"完成：{output_path}")


if __name__ == "__main__":
    main()
